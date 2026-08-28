"""Named short positions from the official EU/UK registers.

This is the only place in the entire section where a SHORT is attributed to a fund
by name. It matters because the 13F structurally cannot show one: it reports long
positions and options, never a short book. A fund can be publicly "long tech" on
its 13F while running a large short against the same theme, and nothing in EDGAR
would ever say so.

The registers publish position holders by name above 0.5% of issued share capital
(UK and EU alike). Below that threshold only aggregates exist.

Hard rule, enforced by the audit: FINRA/exchange short interest and CFTC COT are
ANONYMOUS aggregates. They are context and nothing else. They must never be
written into this table, because presenting an aggregate as a named fund's
position is exactly the kind of quiet fabrication that destroys trust in the whole
platform.
"""

import io
import urllib.error
import urllib.request
from datetime import datetime

from . import db, fund

UA = "Capital Flow research (fund tracker)"
MIN_DISCLOSABLE_PCT = 0.5    # the registers' own reporting threshold
# A register is a log of disclosure EVENTS, not a live position list. A holder that
# last disclosed 1.5% three years ago has almost certainly closed or moved the
# position without a further notification landing in this file — showing it as a
# live short would be inventing a position out of an old announcement. Beyond this
# window the row is kept as history and dropped from the current book.
CURRENT_WINDOW_DAYS = 120


def _fetch(url: str) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    with urllib.request.urlopen(req, timeout=120) as r:
        return r.read()


def parse_fca(blob: bytes) -> list:
    """FCA daily workbook -> rows. Columns: Position Holder, Name of Share Issuer,
    ISIN, Net Short Position (%), Position Date."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    out = []
    for name in wb.sheetnames:
        ws = wb[name]
        header = None
        for row in ws.iter_rows(values_only=True):
            if header is None:
                if row and any(isinstance(c, str) and "position holder" in c.lower()
                               for c in row if c):
                    header = [str(c or "").strip().lower() for c in row]
                continue
            r = dict(zip(header, row))
            holder = r.get("position holder")
            issuer = r.get("name of share issuer")
            if not holder or not issuer:
                continue
            pct = r.get("net short position (%)")
            when = r.get("position date")
            if isinstance(when, datetime):
                when = when.date().isoformat()
            out.append({"fund_name": str(holder).strip(),
                        "issuer": str(issuer).strip(),
                        "isin": (str(r.get("isin") or "").strip() or None),
                        "pct": float(pct) if pct is not None else None,
                        "as_of": str(when)[:10] if when else None,
                        "sheet": name})
    return out


def _match_manager(con, holder: str, index=None):
    """Register name -> tracked manager, or None.

    The registers use legal entity names ("Elliott Advisors (UK) Limited") that do
    not equal the 13F filer name, so matching goes through the same distinctive-
    token test the CIK verification uses, plus the shared alias file. An unmatched
    holder is kept with parent_cik NULL rather than force-fitted onto a tracked
    fund — a wrong attribution here is a fabricated short position.
    """
    if index is None:
        index = [(m["cik"], m["name"]) for m in fund.managers(con)]
    canon = db.resolve_name(holder)
    for cik, name in index:
        if fund._name_agrees(name, canon) or fund._name_agrees(name, holder):
            return cik
    return None


def ingest_register(con, run_id: str, register: str) -> dict:
    cfg = ((fund.load_sources_cfg().get("short_registers") or {}).get(register) or {})
    stats = {"register": register, "rows": 0, "current": 0, "matched": 0,
             "enabled": bool(cfg.get("enabled")), "error": None}
    if not cfg.get("enabled") or not cfg.get("url"):
        # A disabled register is a DECLARED coverage gap, reported as such by the
        # payload. It is never quietly substituted with something weaker.
        fund.log_run(con, run_id, f"shorts:{register}", "skipped",
                     cfg.get("note") or "register disabled in config", stats)
        return stats
    try:
        blob = _fetch(cfg["url"])
        rows = parse_fca(blob) if cfg.get("kind") == "uk" else []
    except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError, OSError,
            ValueError) as exc:
        stats["error"] = str(exc)
        fund.mark_source(con, f"shorts:{register}", ok=False, error=str(exc))
        fund.log_run(con, run_id, f"shorts:{register}", "error", str(exc), stats)
        return stats

    stats["rows"] = len(rows)
    index = [(m["cik"], m["name"]) for m in fund.managers(con)]

    # The register is a disclosure LOG, not a position list: a closed position is
    # reported as a 0% row. So the live book is the latest row per (holder, ISIN),
    # and only rows still at or above the threshold are current.
    latest = {}
    for r in rows:
        if not r["as_of"]:
            continue
        key = (r["fund_name"], r["isin"] or r["issuer"])
        if key not in latest or r["as_of"] > latest[key]["as_of"]:
            latest[key] = r

    cutoff = con.execute("SELECT date('now', ?)",
                         (f"-{CURRENT_WINDOW_DAYS} days",)).fetchone()[0]
    stats["stale_dropped"] = 0
    con.execute("UPDATE fund_shorts SET is_current=0 WHERE register=?", (register,))
    for r in latest.values():
        parent = _match_manager(con, r["fund_name"], index)
        if not parent:
            continue                      # not our universe; nothing to attribute
        fresh = r["as_of"] >= cutoff
        is_current = int((r["pct"] or 0) >= MIN_DISCLOSABLE_PCT and fresh)
        con.execute(
            """INSERT INTO fund_shorts
                 (parent_cik, fund_name, issuer, isin, pct, register, as_of_date,
                  is_current, source_url)
               VALUES (?,?,?,?,?,?,?,?,?)
               ON CONFLICT(register, fund_name, isin, as_of_date) DO UPDATE SET
                 pct=excluded.pct, is_current=excluded.is_current,
                 parent_cik=excluded.parent_cik""",
            (parent, r["fund_name"], r["issuer"], r["isin"], r["pct"], register,
             r["as_of"], is_current, cfg["url"]))
        stats["matched"] += 1
        stats["current"] += is_current
        stats["stale_dropped"] += int((r["pct"] or 0) >= MIN_DISCLOSABLE_PCT and not fresh)
        if is_current:
            fund.add_event(
                con, parent_cik=parent, event_date=r["as_of"],
                disclosed_date=r["as_of"], event_type="short_open",
                headline=(f"{r['pct']:.2f}% net short in {r['issuer']} as of "
                          f"{r['as_of']} — named in the {register} register "
                          f"(a 13F structurally cannot show a short)"),
                issuer=r["issuer"], magnitude=r["pct"],
                magnitude_unit="pct_of_class", source_form=f"{register}-register",
                source_url=cfg["url"])
    con.commit()
    fund.mark_source(con, f"shorts:{register}", ok=True)
    fund.log_run(con, run_id, f"shorts:{register}", "ok",
                 f"{stats['matched']} matched rows, {stats['current']} current", stats)
    return stats


def ingest(con, run_id: str) -> dict:
    regs = (fund.load_sources_cfg().get("short_registers") or {})
    out = {r: ingest_register(con, run_id, r) for r in regs}
    out["declared_gaps"] = [r for r, s in out.items()
                            if isinstance(s, dict) and not s.get("enabled")]
    return out
